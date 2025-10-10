#---------------------------------------------------------------------
# This is to make some changes to the excel files for the MAPPA Areas
"""
1. In sheet "24-25 data", change column headings in columns P,Q and R to Cat1L1caut, Cat1L2caut and Cat1L3caut, respectively.
2. Insert 3 new columns to the right of column R, and these three new columns should have the headings Cat1L1con, Cat1L2con and Cat1L3con, respectively
3. After all these changes have been made, the new colums S,T and U should have the following formulas: ='Data entry'!D121,='Data entry'!D122, and ='Data entry'!D123.
"""
# -------------------------------------------------------------------

import os
import openpyxl
import shutil # to create zip folder

# Change this to your folder path

returns_folder = 'returns'
list_of_file_names = os.listdir(returns_folder)
list_of_file_names = [name for name in list_of_file_names if name.endswith('.xlsx')]
list_of_file_names.remove('NSD.xlsx') # exclude NSD for now. Will work on it separately
len(list_of_file_names) # should be 42 for now without NSD

# Loop through the files and read the specified range
for file_name in list_of_file_names:
    
    # Get the path of the file
    file_path = os.path.join(returns_folder, file_name)
    print(f"Processing {file_path}...")
    
    # Load the file and target sheet to change
    wb = openpyxl.load_workbook(file_path) # the file or workbook
    if "24-25 data" not in wb.sheetnames:
        print(f"  Skipped (no '24-25 data' sheet in {filename})")
        continue
    ws = wb["24-25 data"] # the sheet to change

    # 1. Update headers in P1, Q1, R1
    ws["P1"] = "Cat1L1caut"
    ws["Q1"] = "Cat1L2caut"
    ws["R1"] = "Cat1L3caut"

    # 2. Insert 3 new columns after column R (new S, T, U)
    ws.insert_cols(19, 3)  # 19 = column S (after R)

    # 3. Add new headers
    ws["S1"] = "Cat1L1con"
    ws["T1"] = "Cat1L2con"
    ws["U1"] = "Cat1L3con"

    # 4. Add formulas in row 2
    ws["S2"] = "='Data entry'!D122"
    ws["T2"] = "='Data entry'!D123"
    ws["U2"] = "='Data entry'!D124"

    # Save changes
    wb.save(file_path)

print("✅ All files updated successfully!")



shutil.make_archive("returns_save", 'zip', "returns")
