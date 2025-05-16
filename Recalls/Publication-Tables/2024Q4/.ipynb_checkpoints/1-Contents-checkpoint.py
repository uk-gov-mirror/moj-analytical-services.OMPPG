""" 
GOAL: PRODUCE RECALL TABLES FOR OMSQ.
By Eric Nyame, 17/04/2024
"""

#---------------------------------- Import Packages

import pandas as pd
from pandas.api.types import CategoricalDtype
import numpy as np
import sys
import duckdb
import importlib

# openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

# import re

# from dateutil.relativedelta import relativedelta

# import my predefined functions, akin to macros in SAS

sys.path.append('/home/jovyan/OMPPG/Macro Library')
# from my_log import my_log
import Out_of_bounds_dates
# import prepareMatch
# importlib.reload(prepareMatch)
# import openMatch
# importlib.reload(openMatch)
import TimeDiffs

# Set display options

pd.options.display.max_columns = None
pd.options.display.max_rows = None
pd.set_option('display.max_colwidth', None)

# Ensures no wrapping of cell contents - run it separately

%%html
<style>
.dataframe td {
    white-space: nowrap;
}
</style>


        # Period variables
qtr = 4 # 1:Jan-Mar, 2:Apr-Jun, 3:Jul-Sep, 4:Oct-Dec
year = 2023 # Enter the year being run in 4 digit format

# Bring in recall final datasets----------------------------------------------------------------------
rec1 = pd.read_sas('s3://alpha-omppg/Recalls/final_data/recalls/recalls_final_2022q4.sas7bdat',encoding='latin1')
rec2 = pd.read_sas('s3://alpha-omppg/Recalls/final_data/recalls/recalls_final_2023q1.sas7bdat',encoding='latin1')
rec3 = pd.read_sas('s3://alpha-omppg/Recalls/final_data/recalls/recalls_final_2023q2.sas7bdat',encoding='latin1')
rec4 = pd.read_sas('s3://alpha-omppg/Recalls/final_data/recalls/recalls_final_2023q3.sas7bdat',encoding='latin1')
rec5 = pd.read_parquet(f's3://alpha-omppg/Recalls/final_data/recalls/recalls_final_2023q4.parquet')

# uppercase the headers
for df in [rec1,rec2,rec3,rec4,rec5]:
    df.columns = df.columns.str.upper()

# Concatenate all DataFrames into one------------------------------------------------------------------
recalls = pd.concat([rec1,rec2,rec3,rec4,rec5], ignore_index=True)
len(recalls) # 33912
recalls.head()

# Add a 'QUARTER' column to each DataFrame-----------------------------------------------------------------

# Mapping dictionary
quarter_mapping = {
    'Q1': 'Jan-Mar',
    'Q2': 'Apr-Jun',
    'Q3': 'Jul-Sep',
    'Q4': 'Oct-Dec'
}

# Function to convert period to formatted string
def format_quarter(period):
    year = period.year
    quarter_str = str(period).split('Q')[1]
    return f"{quarter_mapping[f'Q{quarter_str}']} {year}"


# Apply the function
recalls['QUARTER'] = recalls['LICENCE_REVOKE_DATE'].dt.to_period('Q')
recalls['QUARTER'] = recalls['QUARTER'].apply(format_quarter)

# Table 5.2 ------------------------------------------ --------------------------------------------------------

# Change gender values ----------------------------------------------------------------------

gender_mapping = {'F': 'Female', 'M': 'Male'}
recalls['GENDER'] = recalls['GENDER'].replace(gender_mapping)

# Create a SENTENCE column ------------------------------------------------------------------

recalls['SENTENCE'] = np.nan

# Conditions for the 'Sentence' column
conditions = [
    (recalls['CUSTODYTYPE'] == 'Determinate') & (recalls['SENTENCETYPE'].isna()),
    (recalls['CUSTODYTYPE'] == 'Determinate') & (recalls['SENTENCETYPE'] == 'Other'),
    (recalls['CUSTODYTYPE'] == 'Determinate') & (recalls['SENTENCETYPE'] == 'Under 12 months'),
    (recalls['CUSTODYTYPE'] == 'IPP') & (recalls['SENTENCETYPE'].isna()),
    (recalls['CUSTODYTYPE'] == 'IPP') & (recalls['SENTENCETYPE'] == 'Other'),
    (recalls['CUSTODYTYPE'] == 'IPP') & (recalls['SENTENCETYPE'] == 'Under 12 months'),
    (recalls['CUSTODYTYPE'] == 'Life') & (recalls['SENTENCETYPE'].isna()),
    (recalls['CUSTODYTYPE'] == 'Life') & (recalls['SENTENCETYPE'] == 'Other'),
    (recalls['CUSTODYTYPE'] == 'Life') & (recalls['SENTENCETYPE'] == 'Under 12 months')
]

# Corresponding values for each condition
choices = [
    'Missing',
    '12 months or more',
    'Less than 12 months',
    'Missing',
    'IPP',
    'IPP',
    'Missing',
    'Life sentence',
    'Life sentence'
]

# Apply the conditions and choices to 'Sentence'
recalls['SENTENCE'] = np.select(conditions, choices, default=recalls['SENTENCE'])

# Create a HDC indentification ------------------------------------------------------------------

recalls['HDC'] = 'Non-HDC'
recalls.loc[recalls['RECALL_TYPE_DESCRIPTION'].str.contains('HDC', case=False),'HDC'] = 'HDC'


# Define unique values Table 5.2
sex_values = ['Male and female', 'Male', 'Female']
sentence_values = ['All sentence types', 'Less than 12 months', '12 months or more',  'IPP', 'Life sentence']
hdc_values = ['All recalls', 'Non-HDC', 'HDC']
quarters = list(recalls['QUARTER'].unique())

# Function to calculate and append summaries
def calculate_summaries(df, sex_values, sentence_values, hdc_values):
    summary_list = []
    
    for sex in sex_values:
        for sentence in sentence_values:
            if sentence == 'All sentence types':
                temp_df = recalls if sex == 'Male and female' else recalls[recalls['GENDER'] == sex]
            else:
                temp_df = recalls[recalls['SENTENCE'] == sentence] if sex == 'Male and female' else recalls[(recalls['GENDER'] == sex) & (recalls['SENTENCE'] == sentence)]
            
            for quarter in quarters:
                temp_date_df = temp_df[temp_df['QUARTER'] == quarter]
                total_value = temp_date_df['LICENCE_REVOKE_DATE'].count()
                summary_list.append({
                    'Sex': sex,
                    'Sentence type': sentence,
                    'HDC or non-HDC': 'All recalls',
                    'quarter': quarter,
                    'values': total_value
                })
                
                for type in hdc_values[1:]:  # Skip 'All recalls' in this loop
                    type_value = temp_date_df[temp_date_df['HDC'] == type]['LICENCE_REVOKE_DATE'].count()
                    summary_list.append({
                        'Sex': sex,
                        'Sentence type': sentence,
                        'HDC or non-HDC': type,
                        'quarter': quarter,
                        'values': type_value
                    })
    
    return pd.DataFrame(summary_list)


# Calculate summaries for the combined DataFrame
final_summary_df = calculate_summaries(recalls, sex_values, sentence_values, hdc_values)

# order values
final_summary_df['Sex'] = final_summary_df['Sex'].astype(CategoricalDtype(categories=sex_values, ordered=True))
final_summary_df['Sentence type'] = final_summary_df['Sentence type'].astype(CategoricalDtype(categories=sentence_values, ordered=True))
final_summary_df['HDC or non-HDC'] = final_summary_df['HDC or non-HDC'].astype(CategoricalDtype(categories=hdc_values, ordered=True))

# Pivot the final summary DataFrame to get the desired format
final_pivot_df = final_summary_df.pivot_table(
    index=['Sex', 'Sentence type', 'HDC or non-HDC'],
    columns='quarter',
    values='values',
    aggfunc='sum'
).reset_index()


# Rename the columns to match the required format
final_pivot_df = final_pivot_df[['Sex', 'Sentence type', 'HDC or non-HDC'] + quarters]

# percentage change
percent_change_words = f'Percentage change between {quarters[0]} and {quarters[-1]}'
final_pivot_df[percent_change_words] = (final_pivot_df[quarters[-1]] - final_pivot_df[quarters[0]])/ final_pivot_df[quarters[0]]
final_pivot_df[percent_change_words] = np.round(final_pivot_df[percent_change_words],4)

# Format the values to have thousand separators
#for quarter in quarters:
    #final_pivot_df[quarter] = final_pivot_df[quarter].apply(lambda x: f"{int(x):,}" if not pd.isna(x) else x)

output_work_book = f'Tables_{year}Q{qtr}.xlsx'

with pd.ExcelWriter(output_work_book,engine='openpyxl') as writer:
    final_pivot_df.to_excel(writer,startrow = 4, sheet_name='Table 5_12',index=False)
    
# Load the workbook to modify it
workbook = load_workbook(output_work_book)
sheet = workbook['Table 5_12']

# Define the fonts and alignment
title_font = Font(name='Arial', size=16, bold=True)
header_font = Font(name='Arial', size=12, bold=True)
normal_font = Font(name='Arial', size=12)
alignment = Alignment(wrap_text=True)

# Text to be added
header_text = [
    "Table 5_2: Time series: number of recalls from licence, by sex, sentence type, and whether HDC, England and Wales",
    "This worksheet contains one table.",
    "Link to Contents tab",
    "Source: Public Protection Unit Database (PPUD)"
]

# Write the text in the first 4 rows in column A
for i, text in enumerate(header_text):
    cell = sheet.cell(row = i + 1, column=1)
    cell.value = text
    if i == 0:
        cell.font = title_font
    else:
        cell.font = normal_font
    #cell.alignment = Alignment(wrap_text=False)

# Apply font and formatting for the DataFrame cells
for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in row:
        cell.font = normal_font
        # Apply different alignment based on column index
        if cell.column <= 3:  # First three columns
            cell.alignment = Alignment(vertical='bottom',wrap_text=True)
        else:  # Columns 4 to 8
            cell.alignment = Alignment(vertical='bottom', horizontal='right',wrap_text=True)
        cell.border = None

# Format columns 4 to 8 with a thousand separator
for col in range(4, 9):
    for row in range(5, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=col)
        cell.number_format = '#,##0'

# Format the last column as percentages and handle missing values
for row in range(5, sheet.max_row + 1):
    cell = sheet.cell(row=row, column=9)
    if cell.value is None:
        cell.value = '[z]'
        cell.font = Font(name='Arial', size=12)
    else:
        cell.number_format = '0%'
        
# Calculate and set column widths for fixed gaps
def set_fixed_column_widths(sheet, start_row, columns, gap_width):
    for col in columns:
        max_length = 0
        for row in range(start_row, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col)
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + gap_width) 
        col_letter = sheet.cell(row=start_row, column=col).column_letter
        sheet.column_dimensions[col_letter].width = adjusted_width
    
# Set consistent column widths
set_fixed_column_widths(sheet, start_row=5, columns=range(1, 3), gap_width=5)
set_fixed_column_widths(sheet, start_row=5, columns=range(3, 4), gap_width=7)
set_fixed_column_widths(sheet, start_row=5, columns=range(4, 9), gap_width=5)
sheet.column_dimensions['I'].width = 25

# Bold headers (row 5), set row height, and wrap text for the last header cell
for cell in sheet[5]:
    cell.font = header_font
    
sheet.row_dimensions[5].height = 46.8
#sheet.cell(row=5, column=sheet.max_column).alignment = Alignment(wrap_text=True)

# Loop through column 3 to find 'All recalls' and set row height
for row in range(5, sheet.max_row + 1):
    if sheet.cell(row=row, column=3).value == 'All recalls':
        sheet.row_dimensions[row].height = 24.9
        
# Save the modified workbook
workbook.save(output_work_book)
