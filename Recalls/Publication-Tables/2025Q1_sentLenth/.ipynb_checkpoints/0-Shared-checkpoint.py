""" 
GOAL: PRODUCE RECALL TABLES FOR OMSQ.
By Eric Nyame, 17/04/2024
"""

#---------------------------------- Import Packages

import pandas as pd
from pandas.api.types import CategoricalDtype
import numpy as np
import sys
import duckdb
import importlib

from itables import show

# openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Function to identify where bad datetime value is. Pass in the 

def dateOutOfBoundsColumn(dataset,value): # pass in the out-of-bounds date
    for col in dataset.columns:
        # Convert the column to string and check if any value contains the problematic date substring
        if dataset[col].astype(str).str.contains(value).any():
            hmm = dataset[col].astype(str).str.contains(value)
            cols_to_keep = ['FILE_REFERENCE','FAMILY_NAME','LICENCE_REVOKE_DATE',col]
            display(dataset[hmm][cols_to_keep])
            break

# Function to show table in my preferred way
def show_data(data):
    show(data,
         scrollY="200px", 
         scrollCollapse=True, 
         paging=False,
         buttons=["excelHtml5"])
    
# import re

# from dateutil.relativedelta import relativedelta

# import my predefined functions, akin to macros in SAS

sys.path.append('/home/jovyan/OMPPG/Macro-Library')
# from my_log import my_log
import Out_of_bounds_dates
# import prepareMatch
# importlib.reload(prepareMatch)
# import openMatch
# importlib.reload(openMatch)
import TimeDiffs

# Set display options

pd.options.display.max_columns = None
pd.options.display.max_rows = None
pd.set_option('display.max_colwidth', None)


        # Period variables
qtr = 1 # 1:Jan-Mar, 2:Apr-Jun, 3:Jul-Sep, 4:Oct-Dec
year = 2025 # Enter the year being run in 4 digit format

# Quarter mapping dictionary

# Create sentence type function
def sentence(df):
    
    conditions = [
        (df['CUSTODYTYPE'] == 'Determinate') & (df['SENTENCETYPE'].isna()),
        (df['CUSTODYTYPE'] == 'Determinate') & (df['SENTENCETYPE'] == 'Other'),
        (df['CUSTODYTYPE'] == 'Determinate') & (df['SENTENCETYPE'] == 'Under 12 months'),
        (df['CUSTODYTYPE'] == 'IPP') & (df['SENTENCETYPE'].isna()),
        (df['CUSTODYTYPE'] == 'IPP') & (df['SENTENCETYPE'] == 'Other'),
        (df['CUSTODYTYPE'] == 'IPP') & (df['SENTENCETYPE'] == 'Under 12 months'),
        (df['CUSTODYTYPE'] == 'Life') & (df['SENTENCETYPE'].isna()),
        (df['CUSTODYTYPE'] == 'Life') & (df['SENTENCETYPE'] == 'Other'),
        (df['CUSTODYTYPE'] == 'Life') & (df['SENTENCETYPE'] == 'Under 12 months')
    ]

    choices = [
        'Missing',
        'Determinate 12 months or more',
        'Determinate less than 12 months',
        'Missing',
        'IPP',
        'IPP',
        'Missing',
        'Life sentence',
        'Life sentence'
    ]
    
    df['SENTENCE'] = np.nan # set initially to nans
    df['SENTENCE'] = np.select(conditions, choices, default=df['SENTENCE'])

# create ethnicity function

def ethnicity(df):
    table_12_conditions = [
        df['ETHNICITY_DESCRIPTION'].str.contains('mixed', case=False, na=False),
        df['ETHNICITY_DESCRIPTION'].str.contains('black', case=False, na=False),
        df['ETHNICITY_DESCRIPTION'].str.contains('white', case=False, na=False),
        df['ETHNICITY_DESCRIPTION'].str.contains('Asian', case=False, na=False),
        df['ETHNICITY_DESCRIPTION'].str.contains('Refusal', case=False, na=False),
        df['ETHNICITY_DESCRIPTION'].str.contains('chinese', case=False, na=False),
        df['ETHNICITY_DESCRIPTION'].str.contains('Other Ethnic Group', case=False, na=False),
        df['ETHNICITY_DESCRIPTION'].str.contains('Other - Arab', case=False, na=False),
        df['ETHNICITY_DESCRIPTION'].str.contains('Other:', case=False, na=False),
        df['ETHNICITY_DESCRIPTION'].str.contains('Not Known', case=False, na=False),
        df['ETHNICITY_DESCRIPTION'].str.contains('Not Applicable', case=False, na=False),
        df['ETHNICITY_DESCRIPTION'].str.contains('Prefer not to say', case=False, na=False),
        df['ETHNICITY_DESCRIPTION'].str.contains('Any Other', case=False, na=False)
    ]

    choices = [
        'Mixed',
        'Black or Black British',
        'White',
        'Asian or Asian British',
        'Not stated',
        'Asian or Asian British',  # Now classified as Asian
        'Other ethnic group',
        'Other ethnic group',
        'Other ethnic group',
        'Unknown',
        'Unknown',
        'Not stated',
        'Other ethnic group'
    ]

    df['ETHNICITY'] = np.nan # set initially to nans
    df['ETHNICITY'] = np.select(table_12_conditions, choices, default=df['ETHNICITY'])

ethnicity_vals = ['Asian or Asian British', 'Black or Black British','Mixed','White', 'Other ethnic group', 'Not stated']

# Quarter dictionary

quarter_mapping = {
    'Q1': 'Jan to Mar',
    'Q2': 'Apr to June',
    'Q3': 'July to Sept',
    'Q4': 'Oct to Dec'
}

# Create fixed-term/standard term function
def ftrStd(df):
    
    df['RECALL_LENGTH_2'] = 'Standard - determinate sentence'
    df.loc[df['RECALL_TYPE_DESCRIPTION'].str.contains('14',na=False),'RECALL_LENGTH_2'] = 'Fixed term - 14 days'
    df.loc[df['RECALL_TYPE_DESCRIPTION'].str.contains('28',na=False),'RECALL_LENGTH_2'] = 'Fixed term - 28 days'
    df.loc[df['RECALL_TYPE_DESCRIPTION'] == 'Indeterminate Recall','RECALL_LENGTH_2'] = 'Standard - indeterminate sentence'
    df.loc[df['RECALL_TYPE_DESCRIPTION'].str.contains('Emergency',na=False),'RECALL_LENGTH_2'] = 'Other'

# Function to convert period to formatted string for publication

def format_quarter(period):
    year = period.year
    quarter_str = str(period).split('Q')[1] # splits at Q and takes the number at the end
    return f"{quarter_mapping[f'Q{quarter_str}']} {year}"

# Gender mapping

gender_mapping = {'F': 'Female', 'M': 'Male'}

# Define HOWLONG values for Table 8

howlong_mapping = {
    "a Up to and including 6 months": "Up to and including 6 months",
    "b More than 6 months - 1 year": "From 6 months up to and including 12 months",
    "c More than 1 year - 2 years": "From 12 months up to and including 2 years",
    "d More than 2 years - 5 years": "From 2 years up to and including 5 years",
    "e More than 5 years - 10 years": "More than 5 years",
    "f More than 10 years": "More than 5 years",
    "other": "Unknown"
}

howlong_vals = ['All not returned','Up to and including 6 months',
                'From 6 months up to and including 12 months',
                'From 12 months up to and including 2 years',
                'From 2 years up to and including 5 years',
                'More than 5 years']

# Sex, sentence and hdc values for publicationefine unique values

sex_values = ['Male and female', 'Male', 'Female']
sentence_values = ['All sentence types', 'Determinate less than 12 months', 'Determinate 12 months or more',  'IPP', 'Life sentence']
hdc_values = ['All recalls', 'Non-HDC', 'HDC']
ftr_std_values = ['All recalls', 'Fixed term - 14 days', 'Fixed term - 28 days','Standard - determinate sentence','Standard - indeterminate sentence','Other']


# probation_region values = recalls['NPS_CRC_NAME'].unique()

probation_region = ['All regions','East Midlands','East of England','Greater Manchester','Kent Surrey Sussex',
 'London','North East','North West','South Central','South West', 'Wales', 'West Midlands',
 'Yorkshire and The Humber','National Security Division','Unassigned']

supervising_body = ['All supervising bodies', 'National probation', 'Probation trust']

# Define unique values Table 6
rec_process = ['All recall processes', 'Determinate Standard', 'Determinate Emergency', 'Indeterminate Emergency']
return_statuses = ['All return statuses', 'Returned in target', 'Returned outside target', 'Not returned']

# offence groups for Table 9
offence_groups  = ['All offences',
                   'Violence against the person', 
                   'Sexual offences',
                   'Robbery',
                   'Theft offences',
                   'Fraud', 
                   'Drug offences', 
                   'Summary motoring',
                   'Public order offences',
                   'Possession of weapons', 
                   'Miscellaneous crimes against society', 
                   'Criminal damage and arson',  
                   'Offence not recorded']

vatp_subs = ['Homicide', 
             'Death or serious injury caused by illegal driving', 
             'Stalking and harassment',
             'Violence with injury',
             'Violence without injury']

sexual_offence_subs = ['Rape', 'Other sexual offences']

reason_desc_vals = ['Facing further charge',
                    'Non-compliance', 
                    'Failed to keep in touch',
                    'Failed to reside', 
                    'Drugs/alcohol',
                    'Poor Behaviour - Relationships',
                    'HDC - Time violation', 
                    'HDC - Inability to monitor', 
                    'Failed home visit',
                    'HDC - Failed installation', 
                    'HDC - Equipment Tamper',
                    'Other']

# Define the fonts and alignment

title_font = Font(name='Arial', size=16, bold=True)
header_font = Font(name='Arial', size=12, bold=True)
normal_font = Font(name='Arial', size=12)
alignment = Alignment(wrap_text=True)

# Function to set column widths with fixed gaps

def set_fixed_column_widths(sheet, start_row, columns, gap_width):
    for col in columns:
        max_length = 0
        for row in range(start_row, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col)
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + gap_width) 
        col_letter = sheet.cell(row=start_row, column=col).column_letter
        sheet.column_dimensions[col_letter].width = adjusted_width
    
# Function to create table 2

def table_2_func(df, sex_values, sentence_values, hdc_values): # parameters are in Shared.py
    
    # We will create a list of dictionaries of the form 
    # [ {'sex':val, 'Sentence type':val,'Recall category':val, 'quarter': val,'values': val}, ...]
    
    summary_list = []
    
    # Logic:
    # If sex = 'Male and Female' and sentence = 'All sentences', then use all of the dataframe
    # If sex = 'Male and Female' and sentence = '<12', then use df[Sent = '<12'] 
    # If sex = 'Male' and sentence = 'All sentences', then use df['Geder = 'Males']
    # If sex = 'Male' and sentence = '<12', then use df[Gender = 'Males' & Sent ='<12']
    # For each combination of sex and sentence, count entries for each quarter - this is the top line with 'All recalls'
    # For each combination of sex and sentence, filter for true HDC type and count entries for each quarter - this is subsequent line after 'All recalls' line
    
    for sex in sex_values:
        for sentence in sentence_values:
            if sentence == 'All sentence types':
                temp_df = df if sex == 'Male and female' else df[df['GENDER'] == sex]
            else:
                temp_df = df[df['SENTENCE'] == sentence] if sex == 'Male and female' else df[(df['GENDER'] == sex) & (df['SENTENCE'] == sentence)]
            
            # First line for each combination of gender and sentence -> All recall types
            
            for quarter in quarters:
                temp_date_df = temp_df[temp_df['QUARTER'] == quarter]
                total_value = temp_date_df['LICENCE_REVOKE_DATE'].count()
                summary_list.append({
                    'Sex': sex,
                    'Sentence type': sentence,
                    'Recall category': 'All recalls',
                    'quarter': quarter,
                    'values': total_value
                })
                
                # subsequent lines for each combination of gender and senentence - actual HDC types
                
                for type in hdc_values[1:]:  # Skip 'All recalls' in this loop
                    type_value = temp_date_df[temp_date_df['HDC'] == type]['LICENCE_REVOKE_DATE'].count()
                    summary_list.append({
                        'Sex': sex,
                        'Sentence type': sentence,
                        'Recall category': type,
                        'quarter': quarter,
                        'values': type_value
                    })
    # Create a dataframe from the list, where the colums are the keys
    # Sex, Sentence type, Recall category, quarter and values
    
    return pd.DataFrame(summary_list)


# Function to create table 3

def table_3_func(df, probation_region, hdc_values):
    
    summary_list = []
    
    for region in probation_region:
        if region == 'All regions':
            temp_df = df
        else:
            temp_df = df[df['NPS_CRC_NAME'] == region]
            
        for quarter in quarters:
            temp_date_df = temp_df[temp_df['QUARTER'] == quarter]
            total_value = temp_date_df['LICENCE_REVOKE_DATE'].count()
            summary_list.append({
                'Probation region': region,
                'Recall category': 'All recalls',
                'quarter': quarter,
                'values': total_value
                })
            
            for type in hdc_values[1:]:  # Skip 'All recalls' in this loop
                type_value = temp_date_df[temp_date_df['HDC'] == type]['LICENCE_REVOKE_DATE'].count()
                summary_list.append({
                    'Probation region': region,
                    'Recall category': type,
                    'quarter': quarter,
                    'values': type_value
                    })
                    
    return pd.DataFrame(summary_list)

# Function to create table 4

def table_4_func(df, sex_values, sentence_values,table_4_condition):
    
    df = df[table_4_condition]
    
    summary_list = []
    
    for sex in sex_values:
        for sentence in sentence_values:
            if sentence == 'All sentence types':
                temp_df = df if sex == 'Male and female' else df[df['GENDER'] == sex]
            else:
                temp_df = df[df['SENTENCE'] == sentence] if sex == 'Male and female' else df[(df['GENDER'] == sex) & (df['SENTENCE'] == sentence)]
            
            for text in rec_return_by:
                temp_date_df = temp_df[temp_df['RECALLED_IN_RETURNED_BY'] == text]
                total_value = temp_date_df['LICENCE_REVOKE_DATE'].count()
                summary_list.append({
                    'Sex': sex,
                    'Sentence type': sentence,
                    'text': text,
                    'values': total_value
                })
                
    return pd.DataFrame(summary_list)

# Function to calculate Table 5

def table_5_func(df, sex_values, sentence_values,table_5_condition):
    
    df = df[table_5_condition]
    
    summary_list = []
    
    for sex in sex_values:
        for sentence in sentence_values:
            if sentence == 'All sentence types':
                temp_df = df if sex == 'Male and female' else df[df['GENDER'] == sex]
            else:
                temp_df = df[df['SENTENCE'] == sentence] if sex == 'Male and female' else df[(df['GENDER'] == sex) & (df['SENTENCE'] == sentence)]
            
            for text in rec_not_return_by:
                temp_date_df = temp_df[temp_df['RECALLED_IN_NOT_RETURNED_BY'] == text]
                total_value = temp_date_df['LICENCE_REVOKE_DATE'].count()
                summary_list.append({
                    'Sex': sex,
                    'Sentence type': sentence,
                    'text': text,
                    'values': total_value
                })
                
    return pd.DataFrame(summary_list)

# Function to calculate and append summaries

def table_6_func(df, sex_values, sentence_values):
    
    summary_list = []
    
    for process in rec_process:
        for status in return_statuses:
            if status == 'All return statuses':
                temp_df = df if process == 'All recall processes' else df[df['RECALL_PROCESS'] == process]
            else:
                temp_df = df[df['RECALL_TARGET_2'] == status] if process == 'All recall processes' else df[(df['RECALL_PROCESS'] == process) & (df['RECALL_TARGET_2'] == status)]
            
            for text in rec_in_status_on:
                temp_date_df = temp_df[temp_df['RECALLED_IN_STATUS_ON'] == text]
                total_value = temp_date_df['LICENCE_REVOKE_DATE'].count()
                summary_list.append({
                    'Recall process': process,
                    'Return status': status,
                    'text': text,
                    'values': total_value
                })
                
    return pd.DataFrame(summary_list)

def table_7_func(df, sex_values, sentence_values):
    
    summary_list = []
    
    for sex in sex_values:
        for sentence in sentence_values:
            if sentence == 'All sentence types':
                temp_df = df if sex == 'Male and female' else df[df['GENDER'] == sex]
            else:
                temp_df = df[df['SENTENCE'] == sentence] if sex == 'Male and female' else df[(df['GENDER'] == sex) & (df['SENTENCE'] == sentence)]
            
            for text in rec_1984_not_returned_by:
                temp_date_df = temp_df[temp_df['RECALLED_IN_NOT_RETURNED_BY'] == text]
                total_value = temp_date_df['LICENCE_REVOKE_DATE'].count()
                summary_list.append({
                    'Sex': sex,
                    'Sentence type': sentence,
                    'Supervising body': 'All supervising bodies',
                    'text': text,
                    'values': total_value
                })
                
                for body in supervising_body[1:]:  # Skip 'All supervision bodies' in this loop
                    type_value = temp_date_df[temp_date_df['SUP_BODY'] == body]['LICENCE_REVOKE_DATE'].count()
                    summary_list.append({
                        'Sex': sex,
                        'Sentence type': sentence,
                        'Supervising body':body,
                        'text': text,
                        'values': type_value
                    })
                    
    return pd.DataFrame(summary_list)

# Function to create Table 8

def table_8_func(df, supervising_body,howlong_vals):
    
    summary_list = []
    
    for body in supervising_body:
        if body == 'All supervising bodies':
            temp_df = df
        else:
            temp_df = df[df['SUP_BODY'] == body]
            
        for text in rec_1984_not_returned_by:
            temp_date_df = temp_df[temp_df['RECALLED_IN_NOT_RETURNED_BY'] == text]
            total_value = temp_date_df['LICENCE_REVOKE_DATE'].count()
            summary_list.append({
                'Supervising body': body,
                'Time since recall': 'All not returned',
                'text': text,
                'values': total_value
            })
                
            for howlong in howlong_vals[1:]:  # Skip 'All not returned' in this loop
                type_value = temp_date_df[temp_date_df['HOWLONG'] == howlong]['LICENCE_REVOKE_DATE'].count()
                summary_list.append({
                    'Supervising body':body,
                    'Time since recall': howlong,
                    'text': text,
                    'values': type_value
                })
                         
    return pd.DataFrame(summary_list)

# Function to create Table 9

def table_9_func(df, offence_groups,vatp_subs,sexual_offence_subs):
    
    summary_list = []
    
    for offence_group in offence_groups:
        if offence_group == 'All offences':
            temp_df = df
        else:
            temp_df = df[df['OFFENCEGRP_NEW'] == offence_group]
            
        for text in rec_1984_not_returned_by:
            temp_date_df = temp_df[temp_df['RECALLED_IN_NOT_RETURNED_BY'] == text]
            total_value = temp_date_df['LICENCE_REVOKE_DATE'].count()
            summary_list.append({
                'Offence group': offence_group,
                'Offence subgroup': 'All offences',
                'text': text,
                'values': total_value
            })
                
            if offence_group == 'Violence against the person':
                for offence_sub_group in vatp_subs:
                    temp_date_df_2 = temp_date_df[temp_date_df['OFFENCESUBGROUP_NEW'] == offence_sub_group]
                    total_value = temp_date_df_2['LICENCE_REVOKE_DATE'].count()
                    summary_list.append({
                        'Offence group': offence_group,
                        'Offence subgroup': offence_sub_group,
                        'text': text,
                        'values': total_value
                    })
            
            if offence_group == 'Sexual offences':
                for offence_sub_group in sexual_offence_subs:
                    temp_date_df_2 = temp_date_df[temp_date_df['OFFENCESUBGROUP_NEW'] == offence_sub_group]
                    total_value = temp_date_df_2['LICENCE_REVOKE_DATE'].count()
                    summary_list.append({
                        'Offence group': offence_group,
                        'Offence subgroup': offence_sub_group,
                        'text': text,
                        'values': total_value
                    })
                                         
    return pd.DataFrame(summary_list)

def table_10_func(df, df2, sex_values, sentence_values, reason_desc_vals): # parameters are in Shared.py
    
    summary_list = []
    
    for sex in sex_values:
        if sex == 'Male and female':
            for sentence in sentence_values:
                if sentence == 'All sentence types':
                    temp_df = df
                else:
                    temp_df = df[df['SENTENCE'] == sentence]

                # First line for each combination of gender and sentence -> All recall types

                for quarter in quarters:
                    temp_date_df = temp_df[temp_df['QUARTER'] == quarter]              

                    for reason in reason_desc_vals:  # Skip 'All recalls' in this loop
                        reason_value = temp_date_df[temp_date_df['REASON_DESC'] == reason]['LICENCE_REVOKE_DATE'].count()
                        summary_list.append({
                            'Sex': sex,
                            'Sentence type': sentence,
                            'Recall reason': reason,
                            'quarter': quarter,
                            'values': reason_value
                        })
    
        else:
            temp_df = df[df['GENDER'] == sex]
            for quarter in quarters:
                temp_date_df = temp_df[temp_df['QUARTER'] == quarter]              

                for reason in reason_desc_vals:  # Skip 'All recalls' in this loop
                    reason_value = temp_date_df[temp_date_df['REASON_DESC'] == reason]['LICENCE_REVOKE_DATE'].count()
                    summary_list.append({
                            'Sex': sex,
                            'Sentence type': 'All sentence types',
                            'Recall reason': reason,
                            'quarter': quarter,
                            'values': reason_value
                        })
    return pd.DataFrame(summary_list)

def table_12_func(df, sex_values, sentence_values, ethnicity_vals): # parameters are in Shared.py
    
    summary_list = []
    
    for sex in sex_values:
        if sex == 'Male and female':
            for sentence in sentence_values:
                if sentence == 'All sentence types':
                    temp_df = df
                else:
                    temp_df = df[df['SENTENCE'] == sentence]

                # First line for each combination of gender and sentence -> All recall types

                for quarter in quarters:
                    temp_date_df = temp_df[temp_df['QUARTER'] == quarter]              
                    quarter_value = temp_date_df['LICENCE_REVOKE_DATE'].count()
                    summary_list.append({
                            'Sex': sex,
                            'Sentence type': sentence,
                            'Ethnicity': 'All ethnicities',
                            'quarter': quarter,
                            'values': quarter_value
                        })
                    
                    for ethnicity in ethnicity_vals:  # Skip 'All recalls' in this loop
                        ethnicity_value = temp_date_df[temp_date_df['ETHNICITY'] == ethnicity]['LICENCE_REVOKE_DATE'].count()
                        summary_list.append({
                            'Sex': sex,
                            'Sentence type': sentence,
                            'Ethnicity': ethnicity,
                            'quarter': quarter,
                            'values': ethnicity_value
                        })
    
        else:
            temp_df = df[df['GENDER'] == sex]
            for quarter in quarters:
                temp_date_df = temp_df[temp_df['QUARTER'] == quarter]
                quarter_value = temp_date_df['LICENCE_REVOKE_DATE'].count()
                summary_list.append({
                            'Sex': sex,
                            'Sentence type': 'All sentence types',
                            'Ethnicity': 'All ethnicities',
                            'quarter': quarter,
                            'values': quarter_value
                        })
             
                for ethnicity in ethnicity_vals:  # Skip 'All recalls' in this loop
                    ethnicity_value = temp_date_df[temp_date_df['ETHNICITY'] == ethnicity]['LICENCE_REVOKE_DATE'].count()
                    summary_list.append({
                            'Sex': sex,
                            'Sentence type': 'All sentence types',
                            'Ethnicity': ethnicity,
                            'quarter': quarter,
                            'values': ethnicity_value
                        })
    return pd.DataFrame(summary_list)

# Function to create table 13

def table_13_func(df, sex_values, ftr_std_values):
    
    summary_list = []
    
    for sex in sex_values:
        for length in  ftr_std_values:
            if length == 'All recalls':
                temp_df = df if sex == 'Male and female' else df[df['GENDER'] == sex]
            else:
                temp_df = df[df['RECALL_LENGTH_2'] == length] if sex == 'Male and female' else df[(df['GENDER'] == sex) & (df['RECALL_LENGTH_2'] == length)]
            
            for quarter in quarters:
                temp_date_df = temp_df[temp_df['QUARTER'] == quarter]
                total_value = temp_date_df['LICENCE_REVOKE_DATE'].count()
                summary_list.append({
                    'Sex': sex,
                    'Recall length': length,
                    'quarter': quarter,
                    'values': total_value
                })
                
    return pd.DataFrame(summary_list)